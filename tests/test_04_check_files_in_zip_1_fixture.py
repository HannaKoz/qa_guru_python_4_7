import csv
import os
from zipfile import ZipFile

from PyPDF2 import PdfReader
from openpyxl.reader.excel import load_workbook


def test_zip_files():
    with ZipFile('..\\resources\\myZip.zip') as my_zip:
        files_list = list(my_zip.namelist())
        print(files_list)
        for files in my_zip.infolist():
            print(f"{files.filename},\t {files.compress_type},\t {files.file_size},\t {files.compress_size}")
        assert 'dummy.pdf' in files_list
        assert 'username.csv' in files_list
        assert 'Financial Sample.xlsx' in files_list


def test_check_pdf():
    with ZipFile('..\\resources\\myZip.zip') as my_zip:
        with my_zip.open('dummy.pdf') as pdf_test:
            reader = PdfReader(pdf_test, 'r')
            number_of_pages = len(reader.pages)
            print(f"number of pages: {number_of_pages}")
            assert number_of_pages is 155
            text = reader.pages[144].extract_text().__contains__('Copyright (c) 1991, 2000, 2001 by Lucent '
                                                                 'Technologies.')
            assert text is True
            text_len = len(reader.pages[6].extract_text())
            print(text_len)
            assert text_len == 1983


def test_check_xlsx():
    with ZipFile('..\\resources\\myZip.zip') as my_zip:
        with my_zip.open('Financial Sample.xlsx') as xlsx_test:
            book = load_workbook(xlsx_test)
            sheets = book.active
            cell_value = sheets.cell(row=3, column=2).value
            print(cell_value)
            assert cell_value == 'Germany'
            for x_file in my_zip.infolist():
                name = os.path.basename(x_file.filename)
                if name == 'Financial Sample.xlsx':
                    size = x_file.file_size
                    print(f"{name},\t {size}")
            assert size == 83418


def test_check_csv():
    with ZipFile('..\\resources\\myZip.zip') as my_zip:
        with my_zip.open('username.csv') as csv_test:
            csv_file = csv.reader(csv_test)
            print(csv_file)
            for csv_file in my_zip.infolist():
                csv_name = os.path.basename(csv_file.filename)
                if csv_name == 'username.csv':
                    size_csv = csv_file.file_size
                    print(f"File name: {csv_name},\t size:{size_csv}")
            assert size_csv == 176


def test_delete_files():
    # remove_xlsx:
    os.remove('..\\tests\\Financial Sample.xlsx')
    # remove_pdf:
    os.remove('..\\resources\\dummy.pdf')
    # remove_csv:
    os.remove('..\\resources\\username.csv')
    # remove_zip
    os.remove('..\\resources\\myZip.zip')
